import os
import sys
from pathlib import Path
import re
import pandas as pd
from datetime import datetime
from PIL import Image
import easyocr
import fitz  # PyMuPDF
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
import time
import threading

# Initialize EasyOCR reader
reader = easyocr.Reader(['en'])

def setup_environment():
    """Setup and validate the environment."""
    print("Current working directory:", os.getcwd())

def read_chat_file(filepath):
    """Read the WhatsApp chat file."""
    print(f"Reading chat file: {filepath}")
    with open(filepath, 'r', encoding='utf-8') as file:
        return file.read()

def extract_data_from_text(text):
    """Extract data from the chat text using regex."""
    print("Extracting data from chat text")
    pattern = r"(\d{1,2}/\d{1,2}/\d{2,4}), (\d{1,2}:\d{2}\s?[APMapm]*) - ([^:]+): (.*)"
    return re.findall(pattern, text)

def extract_text_from_pdf(pdf_path):
    """Extract text from a PDF file."""
    text = ""
    try:
        pdf_document = fitz.open(pdf_path)
        for page_num in range(len(pdf_document)):
            page = pdf_document.load_page(page_num)
            text += page.get_text()
    except Exception as e:
        print(f"Error reading PDF {pdf_path}: {e}")
    return text

def process_message(date_str, time_str, name, message, start_date, end_date):
    """Process each message to extract relevant information."""
    try:
        message_date = datetime.strptime(date_str, "%m/%d/%Y")
    except ValueError:
        try:
            message_date = datetime.strptime(date_str, "%m/%d/%y")
        except ValueError:
            print(f"Error parsing date: {date_str}")
            return None
    
    if not (start_date <= message_date <= end_date):
        return None

    number = 0
    entry_type = "Mensaje"
    image_path = ""
    
    if "IMG-" in message or "Comprobante_" in message:
        entry_type = "Imagen"
        image_path = os.path.join('expenses/data', message.split(' ')[0])
        if os.path.exists(image_path):
            try:
                if image_path.lower().endswith('.pdf'):
                    ocr_text = extract_text_from_pdf(image_path)
                else:
                    img = Image.open(image_path)
                    ocr_text = reader.readtext(image_path, detail=0)
                    ocr_text = ' '.join(ocr_text)
                message += f" OCR: {ocr_text} (Image path: {image_path})"
                ocr_numbers = re.findall(r"\$\s?(\d{1,3}(?:\.\d{3})*(?:,\d{2})?)", ocr_text)
                if ocr_numbers:
                    number = float(ocr_numbers[0].replace('.', '').replace(',', '.'))
                else:
                    number = "Verifique"
            except Exception as e:
                print(f"Error processing image {image_path}: {e}")
                number = "Verifique"
    elif "STK-" in message or "PTT-" in message:
        return None
    else:
        number_match = re.search(r"(\d+(?:\.\d{2})?)", message)
        if number_match:
            number = float(number_match.group(1).replace('.', '').replace(',', '.'))
    
    return {"Fecha": date_str, "Hora": time_str, "Nombre": name, "Mensaje": message, "Monto": number, "Tipo": entry_type, "Path": image_path}

def create_dataframe(matches, start_date, end_date):
    """Create a DataFrame from the extracted data."""
    print("Creating DataFrame from extracted data")
    data = []
    for m in matches:
        result = process_message(m[0], m[1], m[2], m[3], start_date, end_date)
        if result:
            data.append(result)
    return pd.DataFrame(data)

def save_to_excel(df, summary, excel_path):
    """Save the DataFrame and summary to an Excel file."""
    print(f"Saving data to Excel file: {excel_path}")
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Detalles', index=False)
        summary.to_excel(writer, sheet_name='Resumen', index=False)

def format_excel(excel_path):
    """Format the Excel file."""
    print("Formatting Excel file")
    wb = load_workbook(excel_path)
    ws_detalles = wb['Detalles']
    ws_resumen = wb['Resumen']

    table = Table(displayName="DetallesTable", ref=ws_detalles.dimensions)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style
    ws_detalles.add_table(table)

    ws_detalles.column_dimensions['D'].width = 50
    for row in ws_detalles.iter_rows(min_row=2, max_row=ws_detalles.max_row):
        ws_detalles.row_dimensions[row[0].row].height = 30

    wb.save(excel_path)

def process_data(start_date, end_date):
    """Process the data and create the Excel report."""
    # Find WhatsApp chat file in the data directory
    print("Searching for WhatsApp chat file in expenses/data directory")
    chat_files = list(Path('expenses/data').glob('WhatsApp Chat*.txt'))
    if not chat_files:
        print("No WhatsApp chat file found in expenses/data directory")
        return
    
    start_time = time.time()
    text = read_chat_file(str(chat_files[0]))
    matches = extract_data_from_text(text)
    df = create_dataframe(matches, start_date, end_date)
    print(f"Data extraction and processing took {time.time() - start_time:.2f} seconds")
    
    start_time = time.time()
    df['Monto'] = pd.to_numeric(df['Monto'], errors='coerce').fillna(0)
    summary = df.groupby('Nombre').agg({'Monto': 'sum', 'Mensaje': lambda x: (x == "Verifique").sum()}).reset_index()
    summary.columns = ['Nombre', 'Total Monto', 'Verifique Count']
    print(f"Data summarization took {time.time() - start_time:.2f} seconds")
    
    excel_path = 'expenses/chat_whatsapp.xlsx'
    start_time = time.time()
    save_to_excel(df, summary, excel_path)
    format_excel(excel_path)
    print(f"Excel file creation and formatting took {time.time() - start_time:.2f} seconds")
    
    try:
        if os.path.exists(excel_path):
            print(f"Opening Excel file: {excel_path}")
            os.startfile(os.path.join(os.getcwd(), excel_path))
        else:
            print(f"Error: The file {excel_path} was not created")
    except Exception as e:
        print(f"Error opening the Excel file: {e}")

def main():
    setup_environment()
    
    start_date_str = input("Ingrese la fecha de inicio (dd/mm/yyyy): ")
    end_date_str = input("Ingrese la fecha de fin (dd/mm/yyyy): ")
    
    try:
        start_date = datetime.strptime(start_date_str, "%d/%m/%Y")
        end_date = datetime.strptime(end_date_str, "%d/%m/%Y")
    except ValueError:
        print("Formato de fecha incorrecto. Use dd/mm/yyyy.")
        return
    
    # Run data processing in a separate thread
    processing_thread = threading.Thread(target=process_data, args=(start_date, end_date))
    processing_thread.start()
    
    # Display progress message
    while processing_thread.is_alive():
        print("Creando el reporte...", end="\r")
        time.sleep(1)
    
    print("Reporte creado exitosamente.")

if __name__ == "__main__":
    main()



